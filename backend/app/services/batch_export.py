"""Tabular Excel export for batch data processing results.

Generates a clean, analysis-ready Excel workbook (not the report-style
format used by risk/sentiment/tech agents).
"""

import io
import csv
from typing import Any


def generate_batch_excel(result: dict, agent_type: str = "batch") -> bytes:
    """Generate a tabular Excel workbook from batch processing result.

    Sheet 1 "数据明细"  — full dataset with auto-filter and frozen header
    Sheet 2 "统计摘要"  — descriptive statistics and query metadata
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    AGENT_COLOR = "06B6D4"   # cyan — batch agent accent
    HEADER_TEXT = "FFFFFF"
    ALT_ROW_COLOR = "F0FAFA"

    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: 数据明细
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "数据明细"

    columns: list[dict] = result.get("columns", [])
    # Use full_rows from metadata if available, else preview_rows
    metadata = result.get("metadata", {})
    full_rows: list[dict[str, Any]] = metadata.get("full_rows", [])
    if not full_rows:
        full_rows = result.get("preview_rows", [])

    # Header row
    header_font = Font(bold=True, color=HEADER_TEXT, name="Microsoft YaHei", size=10)
    header_fill = PatternFill(start_color=AGENT_COLOR, end_color=AGENT_COLOR, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )

    for col_idx, col in enumerate(columns, 1):
        label = col.get("label", col.get("key", ""))
        unit = col.get("unit", "")
        header_text = f"{label}\n({unit})" if unit else label
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 32

    # Data rows
    alt_fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")
    data_font = Font(name="Microsoft YaHei", size=9)
    data_align_left = Alignment(horizontal="left", vertical="center")
    data_align_right = Alignment(horizontal="right", vertical="center")

    BADGE_LABELS = {
        "listed": "上市",
        "unlisted": "未上市",
        "low": "低风险",
        "medium": "中风险",
        "high": "高风险",
    }

    for row_idx, row_data in enumerate(full_rows, 2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, col in enumerate(columns, 1):
            key = col.get("key", "")
            col_type = col.get("type", "text")
            raw_val = row_data.get(key)

            # Convert badge codes to human labels
            if col_type == "badge" and isinstance(raw_val, str):
                value = BADGE_LABELS.get(raw_val, raw_val)
            elif raw_val is None:
                value = ""
            else:
                value = raw_val

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill
            cell.alignment = (
                data_align_right if col_type == "number" else data_align_left
            )

    # Auto-filter on header row
    if columns:
        last_col_letter = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col_letter}1"
        ws.freeze_panes = "A2"

    # Auto-fit column widths (heuristic)
    for col_idx, col in enumerate(columns, 1):
        label = col.get("label", "")
        max_len = len(label) + 4
        for row_data in full_rows[:50]:  # sample first 50 rows
            val = row_data.get(col.get("key", ""), "")
            max_len = max(max_len, len(str(val) if val is not None else ""))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len * 1.8 + 2, 38)

    # ------------------------------------------------------------------
    # Sheet 2: 统计摘要
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("统计摘要")

    title_font = Font(bold=True, color=AGENT_COLOR, name="Microsoft YaHei", size=12)
    label_font = Font(bold=True, name="Microsoft YaHei", size=9, color="555555")
    value_font = Font(name="Microsoft YaHei", size=9)

    stats = result.get("stats", {})
    query = metadata.get("query", result.get("sql_description", "批量数据查询"))
    scenario_map = {"filter": "筛选并导出企业名单", "export": "指定字段结构化导出", "derived": "衍生加工字段导出"}
    scenario = scenario_map.get(metadata.get("scenario", "filter"), "筛选并导出企业名单")
    total_count = result.get("total_count", len(full_rows))
    duration = metadata.get("duration", "--")

    rows_info = [
        ("查询报告", "批量数据处理结果"),
        ("", ""),
        ("查询需求", query),
        ("处理场景", scenario),
        ("匹配企业数", f"{total_count} 家"),
        ("执行耗时", f"{duration}s"),
        ("", ""),
        ("— 统计摘要 —", ""),
    ]

    # Add stats
    for key, val in stats.items():
        if key == "count":
            continue
        # Convert snake_case key to human label
        label = key.replace("avg_", "平均 ").replace("max_", "最大 ").replace(
            "min_", "最小 ").replace("median_", "中位数 ").replace("_", " ").title()
        rows_info.append((label, f"{val:,.1f}" if isinstance(val, float) else str(val)))

    dq = result.get("data_quality", {})
    if dq:
        rows_info += [
            ("", ""),
            ("— 数据质量 —", ""),
            ("请求记录数", str(dq.get("total_requested", total_count))),
            ("匹配成功", str(dq.get("matched", total_count))),
            ("缺失值补全", str(dq.get("missing_fields_filled", 0))),
        ]

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 40

    for i, (lbl, val) in enumerate(rows_info, 1):
        cell_a = ws2.cell(row=i, column=1, value=lbl)
        cell_b = ws2.cell(row=i, column=2, value=val)
        if lbl in ("查询报告", "— 统计摘要 —", "— 数据质量 —"):
            cell_a.font = title_font
        elif lbl:
            cell_a.font = label_font
            cell_b.font = value_font
        else:
            cell_a.font = value_font
            cell_b.font = value_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_batch_csv(result: dict) -> bytes:
    """Generate CSV bytes as a fallback when Excel dependencies are unavailable."""
    columns: list[dict] = result.get("columns", [])
    metadata = result.get("metadata", {})
    full_rows: list[dict[str, Any]] = metadata.get("full_rows", [])
    if not full_rows:
        full_rows = result.get("preview_rows", [])

    output = io.StringIO()
    writer = csv.writer(output)

    if columns:
        writer.writerow([col.get("label", col.get("key", "")) for col in columns])
        for row in full_rows:
            writer.writerow([row.get(col.get("key", ""), "") for col in columns])
    elif full_rows:
        keys = list(full_rows[0].keys())
        writer.writerow(keys)
        for row in full_rows:
            writer.writerow([row.get(key, "") for key in keys])
    else:
        writer.writerow(["message"])
        writer.writerow(["No data available"])

    # UTF-8 BOM for Excel compatibility.
    return ("\ufeff" + output.getvalue()).encode("utf-8")

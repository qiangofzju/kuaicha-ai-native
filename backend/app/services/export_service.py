"""Export service – generate PDF and Excel files from agent results."""

import io
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

AGENT_LABELS = {
    "risk": "企业风控尽调报告",
    "sentiment": "舆情风险简报",
    "tech": "科创能力评估报告",
}

RISK_LEVEL_CN = {"high": "高", "medium": "中", "low": "低"}

# Try to register a CJK font; fall back to Helvetica if unavailable
_CJK_FONT = "Helvetica"
try:
    pdfmetrics.registerFont(TTFont("NotoSansSC", "/System/Library/Fonts/PingFang.ttc", subfontIndex=0))
    _CJK_FONT = "NotoSansSC"
except Exception:
    try:
        pdfmetrics.registerFont(TTFont("NotoSansSC", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", subfontIndex=0))
        _CJK_FONT = "NotoSansSC"
    except Exception:
        pass  # Fallback to Helvetica – CJK chars may not render


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def generate_pdf(result: dict, agent_type: str) -> bytes:
    """Generate a PDF report from an agent result dict."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=20 * mm, rightMargin=20 * mm, topMargin=20 * mm, bottomMargin=20 * mm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("CJKTitle", parent=styles["Title"], fontName=_CJK_FONT, fontSize=18, spaceAfter=12))
    styles.add(ParagraphStyle("CJKHeading", parent=styles["Heading2"], fontName=_CJK_FONT, fontSize=14, spaceAfter=8, spaceBefore=14))
    styles.add(ParagraphStyle("CJKBody", parent=styles["BodyText"], fontName=_CJK_FONT, fontSize=10, leading=16, spaceAfter=6))
    styles.add(ParagraphStyle("CJKSmall", parent=styles["BodyText"], fontName=_CJK_FONT, fontSize=8, leading=12, textColor=HexColor("#666666")))

    story = []

    # Title
    title = AGENT_LABELS.get(agent_type, "分析报告")
    story.append(Paragraph(title, styles["CJKTitle"]))

    target = result.get("metadata", {}).get("target", "")
    if target:
        story.append(Paragraph(f"目标企业：{target}", styles["CJKBody"]))
    story.append(Paragraph(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["CJKSmall"]))
    story.append(Spacer(1, 8 * mm))

    # Summary
    summary = result.get("summary", "")
    if summary:
        story.append(Paragraph("综合摘要", styles["CJKHeading"]))
        story.append(Paragraph(summary, styles["CJKBody"]))
        story.append(Spacer(1, 4 * mm))

    # Risk rating
    rating = result.get("risk_rating", "")
    if rating:
        story.append(Paragraph(f"综合评级：{rating}", styles["CJKBody"]))
        story.append(Spacer(1, 4 * mm))

    # Findings table
    findings = result.get("risk_findings", [])
    if findings:
        story.append(Paragraph("关键发现", styles["CJKHeading"]))
        table_data: list[list[str]] = [["维度", "等级", "评分", "说明"]]
        for f in findings:
            level = f.get("level", "")
            table_data.append([
                f.get("label", "") or "",
                RISK_LEVEL_CN.get(level, level) if level else "",
                str(f.get("score", "")),
                f.get("description", "") or "",
            ])

        t = Table(table_data, colWidths=[60, 40, 40, 320])
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), _CJK_FONT),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#f0f0f0")),
            ("GRID", (0, 0), (-1, -1), 0.5, HexColor("#dddddd")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 4 * mm))

    # Report sections
    sections = result.get("report_sections", [])
    for sec in sections:
        story.append(Paragraph(sec.get("title", ""), styles["CJKHeading"]))
        story.append(Paragraph(sec.get("content", ""), styles["CJKBody"]))

    # Timeline
    timeline = result.get("timeline", [])
    if timeline:
        story.append(Paragraph("时间线", styles["CJKHeading"]))
        for item in timeline:
            story.append(Paragraph(
                f"[{item.get('date', '')}] {item.get('event', '')}",
                styles["CJKBody"],
            ))

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def generate_excel(result: dict, agent_type: str) -> bytes:
    """Generate an Excel workbook from an agent result dict."""
    wb = Workbook()

    header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4A9EFF", end_color="4A9EFF", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    # --- Sheet 1: Summary ---
    ws = wb.active  # type: ignore[assignment]
    ws.title = "概要"  # type: ignore[union-attr]
    ws.column_dimensions["A"].width = 16  # type: ignore[union-attr]
    ws.column_dimensions["B"].width = 60  # type: ignore[union-attr]

    title = AGENT_LABELS.get(agent_type, "分析报告")
    target = result.get("metadata", {}).get("target", "")

    rows = [
        ("报告类型", title),
        ("目标企业", target),
        ("综合评级", result.get("risk_rating", "N/A")),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("综合摘要", result.get("summary", "")),
    ]
    for row in rows:
        ws.append(row)  # type: ignore[union-attr]

    # --- Sheet 2: Findings ---
    findings = result.get("risk_findings", [])
    if findings:
        ws2 = wb.create_sheet("关键发现")
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 10
        ws2.column_dimensions["D"].width = 60

        headers = ["维度", "等级", "评分", "说明"]
        ws2.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for f in findings:
            ws2.append([
                f.get("label", ""),
                RISK_LEVEL_CN.get(f.get("level", ""), f.get("level", "")),
                f.get("score", ""),
                f.get("description", ""),
            ])

    # --- Sheet 3: Report ---
    sections = result.get("report_sections", [])
    if sections:
        ws3 = wb.create_sheet("详细报告")
        ws3.column_dimensions["A"].width = 20
        ws3.column_dimensions["B"].width = 80

        headers = ["章节", "内容"]
        ws3.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for sec in sections:
            ws3.append([sec.get("title", ""), sec.get("content", "")])

    # --- Sheet 4: Timeline ---
    timeline = result.get("timeline", [])
    if timeline:
        ws4 = wb.create_sheet("时间线")
        ws4.column_dimensions["A"].width = 14
        ws4.column_dimensions["B"].width = 60
        ws4.column_dimensions["C"].width = 10

        headers = ["日期", "事件", "等级"]
        ws4.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws4.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for item in timeline:
            ws4.append([
                item.get("date", ""),
                item.get("event", ""),
                item.get("level", ""),
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
